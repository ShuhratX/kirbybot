"""
report_excel.py — buyurtmalar ro'yxatini .xlsx faylga chiqarish.
openpyxl kutubxonasi kerak: pip install openpyxl
"""
from __future__ import annotations
import json
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def build_report(rows: list[dict], date_from: str, date_to: str) -> bytes:
    """
    rows — database.report_by_range() dan kelgan list[dict].
    Qaytaradi: xlsx fayl bytes (bot.py da BufferedInputFile uchun).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hisobot"

    # ── Sarlavha ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Buyurtmalar hisoboti: {date_from} — {date_to}"
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = Alignment(horizontal="center")

    # ── Ustun sarlavhalari ────────────────────────────────────────────────────
    headers = ["№", "Mijoz ismi", "Telefon", "Mahsulotlar", "Buyurtma summasi", "Sana"]
    header_fill = PatternFill("solid", fgColor="0EA5E9")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # ── Ma'lumotlar ──────────────────────────────────────────────────────────
    grand_total = 0.0

    for i, order in enumerate(rows, start=1):
        row_num = i + 2

        # items JSON ni o'qiymiz
        try:
            items: list[dict] = json.loads(order["items"] or "[]")
        except (json.JSONDecodeError, TypeError):
            items = []

        # "Mahsulot A x2, Mahsulot B x1" ko'rinishida
        items_str = ", ".join(
            f"{item.get('name', '?')} x{item.get('qty', 1)}"
            for item in items
        ) or "—"

        total = float(order["total"] or 0)
        grand_total += total

        row_data = [
            i,
            order["full_name"] or "—",
            order["phone"] or "—",
            items_str,
            total,
            (order["created_at"] or "")[:10],  # faqat sana
        ]

        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=(col == 4))
            cell.border = border
            if col == 5:  # summa ustuni
                cell.number_format = '#,##0 "so\'m"'

    # ── Jami summa qatori ────────────────────────────────────────────────────
    total_row = len(rows) + 3
    ws.merge_cells(f"A{total_row}:D{total_row}")
    total_label = ws[f"A{total_row}"]
    total_label.value = "JAMI:"
    total_label.font = Font(bold=True, size=11)
    total_label.alignment = Alignment(horizontal="right")

    total_val = ws.cell(row=total_row, column=5, value=grand_total)
    total_val.font = Font(bold=True, size=11)
    total_val.number_format = '#,##0 "so\'m"'
    total_val.fill = PatternFill("solid", fgColor="DCFCE7")
    total_val.border = border

    # ── Ustun kengliklari ────────────────────────────────────────────────────
    col_widths = [5, 22, 16, 40, 18, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[2].height = 20

    # ── Bytes sifatida qaytarish ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()